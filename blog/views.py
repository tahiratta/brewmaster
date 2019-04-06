from django.shortcuts import render, get_object_or_404, redirect, HttpResponseRedirect, reverse
# from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
# from django.contrib.auth.models import User
# from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from .models import *
from custdata.models import *
from .forms import *
from django.contrib.auth.decorators import login_required
import openpyxl
from django.contrib import messages
import logging
from datetime import datetime, timedelta
from django.db.models import Q
import decimal
import xlsxwriter
from django_project.settings import MEDIA_URL


# Create your views here.
@login_required
def home(request):
    if Customers.objects.all().exists():

        if request.method == 'POST':
            print('this',request.POST.get('count'))

            fromDate = request.POST.get('fromDate')
            print('from date check kro zara', fromDate)
            first = datetime.strptime('1900-01-01', '%Y-%m-%d')
            if fromDate is not '':
                first = datetime.strptime(fromDate, '%Y-%m-%d')
                print('pehli date',first)
                # first_ago = first - timedelta(days=1)

            toDate = request.POST.get('toDate')
            last_advent = datetime.strptime('1900-01-01', '%Y-%m-%d')
            if toDate is not '':
                last = datetime.strptime(toDate, '%Y-%m-%d')
                last_advent = last #+ timedelta(days=1)
                print('aakhri date', last_advent)

            greater = datetime.strptime('2050-01-01', '%Y-%m-%d')
            gDate = request.POST.get('gDate')
            if gDate is not '':
                greater = datetime.strptime(gDate, '%Y-%m-%d')

            lower = datetime.strptime('1900-01-01', '%Y-%m-%d')
            lDate = request.POST.get('lDate')
            if lDate is not '':
                lower = datetime.strptime(lDate, '%Y-%m-%d')

            order_total = request.POST.get('total')
            order_count = request.POST.get('count')

            order_total = -1 if order_total == '' else float(order_total)
            order_count = -1 if order_count == '' else int(order_count)

            print('totla and count', order_total, order_count)

            date_param = Q(order_date__range=[first, last_advent])
            ldate_param = Q(order_date__lt=lower)
            gdate_param = Q(order_date__gt=greater)
            # order_total1 = Q(order_total__gt=order_total)

            # print('Q hai ye', order_total1)

            customers_list = []
            ord_total_cust = []
            all_customers = Customers.objects.all()
            order_customer_id = Q(customer_id=0)
            order_total_cust_id = Q(customer_id=0)
            for c in all_customers:
                customer_key = c.customer_id
                all_orders = Orders.objects.filter(customer_id=customer_key)
                count = 0
                sum = 0.00
                refund = 0.00
                for o in all_orders:
                    customer_id = o.customer_id.customer_id
                    order_total_value = o.order_total
                    sum = sum + float(order_total_value)
                    order_total_refund = o.refund_amount
                    refund = refund + float(order_total_refund)
                    count += 1

                total_sum = sum - refund
                print('find total sum',total_sum)
                print('is that true or false',total_sum == order_total)
                if total_sum > order_total and order_total != -1:
                    print('aa gya is k andar bhi bhai jaan')
                    ord_total_cust.append(customer_id)
                    order_total_cust_id = Q(customer_id__in=ord_total_cust)

                if count > order_count and order_count != -1:
                    print('aa gya is k andar')
                    customers_list.append(customer_id)
                    order_customer_id = Q(customer_id__in=customers_list)

            print('saare params ajain yahan',date_param, ldate_param, gdate_param, order_total_cust_id, order_customer_id, order_count, count)

            print('i want the select *',fromDate, toDate, gDate, lDate, order_total, order_count)

            if fromDate == '' and toDate == '' and gDate == '' and lDate == '' and order_total == -1 and order_count == -1:
                print('if part time')
                orders = Orders.objects.all().values_list('customer_id__first_name', 'customer_id__last_name',
                'customer_id__address_1', 'customer_id__address_2', 'customer_id__city', 'customer_id__state',
                'customer_id__zip_code', 'customer_id__country', 'customer_id__country_code').distinct()
            else:
                orders = Orders.objects.order_by("-order_date").filter(
                    date_param | ldate_param | gdate_param | order_total_cust_id | order_customer_id).values_list('customer_id__first_name', 'customer_id__last_name',
                'customer_id__address_1', 'customer_id__address_2', 'customer_id__city', 'customer_id__state',
                'customer_id__zip_code', 'customer_id__country', 'customer_id__country_code').distinct()


                print('else part time')
                # print('going it or nt',Orders.objects.order_by("-order_date").filter(order_customer_id))
                print(orders)

            if orders:
                columns = request.POST.getlist('dropdown')
                read = []
                l = 0
                for ord in orders:
                    for i in columns:
                        print('i my meS', i)
                        if i == 'first_name':
                            print(ord[0])
                            j = ord[0]

                        if i == 'last_name':
                            j = ord[1]

                        if i == 'address_1':
                            j = ord[2]

                        if i == 'address_2':
                            j = ord[3]

                        if i == 'city':
                            j = ord[4]

                        if i == 'state':
                            j = ord[5]

                        if i == 'zip_code':
                            j = ord[6]

                        if i == 'country':
                            j = ord[7]

                        if i == 'country_code':
                            j = ord[8]

                        if j != '':
                            read.append(j)
                            print('Please write it', read)
                    l += 1


                # # s = orders[0].order_date.strftime('%Y%m%d%H%M%S%f')
                # s = datetime.now().strftime('%Y%m%d%H%M%S%f')
                # print(s)
                # tail = s[-4:]
                # f = round(float(tail), 3)
                # temp = "%.3f" % f
                # print((s[:-4], temp[1:]))
                # print(type(s[-4:]))
                # date_time = s[:-4]
                #
                # filename = "orders" + date_time + ".xlsx"
                # wb = xlsxwriter.Workbook(filename)
                # ws = wb.add_worksheet("responses %s" % date_time)
                # row_num = 0
                #
                # # font_style = xlsxwriter.XFStyle()
                # font_style = wb.add_format()
                # font_style.set_bold()
                # # font_style.font.bold = True
                #
                # print('dropdown bhai jaan', request.POST.getlist('dropdown'))
                #
                # columns = request.POST.getlist('dropdown')
                # # columns = ['first_name', 'last_name', 'address_1', 'address_2', 'city', 'state', 'zip_code', 'country', 'country_code' ]
                #
                # for col_num in range(len(columns)):
                #     ws.write(row_num, col_num, columns[col_num], font_style)
                #
                # font_style = wb.add_format()
                # font_style.set_bold(False)
                #
                # write = []
                # k = 0
                # for ord in orders:
                #     for i in columns:
                #         if row_num <= 1048574:
                #             print('i my meS', i)
                #             if i == 'first_name':
                #                 j = ord.customer_id.first_name
                #
                #             if i == 'last_name':
                #                 j = ord.customer_id.last_name
                #
                #             if i == 'address_1':
                #                 j = ord.customer_id.address_1
                #
                #             if i == 'address_2':
                #                 j = ord.customer_id.address_2
                #
                #             if i == 'city':
                #                 j = ord.customer_id.city
                #
                #             if i == 'state':
                #                 j = ord.customer_id.state
                #
                #             if i == 'zip_code':
                #                 j = ord.customer_id.zip_code
                #
                #             if i == 'country':
                #                 j = ord.customer_id.country
                #
                #             if i == 'country_code':
                #                 j = ord.customer_id.country_code
                #
                #         if j != '':
                #             write.append(j)
                #             print('Please write it', write)
                #
                #     # print("before increment")
                #     # print(row_num)
                #     # row_num += 1
                #     # print("after increment")
                #     # print(row_num)
                #     # # print(row)
                #     # ws.write(row_num, 0, write, font_style)
                #     # ws.write(row_num, 1, write, font_style)
                #     # ws.write(row_num, 2, write, font_style)
                #     # ws.write(row_num, 3, write, font_style)
                #     # ws.write(row_num, 4, write, font_style)
                #     # ws.write(row_num, 5, write, font_style)
                #     # ws.write(row_num, 6, write, font_style)
                #     # ws.write(row_num, 7, write, font_style)
                #     # ws.write(row_num, 8, write, font_style)
                #     # print('ab tak row num', row_num)
                #
                #     k += 1
                #     for row_num in range(len(write)):
                #         ws.write(k, row_num, write[row_num], font_style)
                #
                # row_num += 2
                # wb.close()
                # # return Response({"status": status.HTTP_200_OK, "link": "Downloads/" + filename})
                # print('success')
                msg=''
                col_length = len(columns)
                print('length of cols', col_length)
                return render(request, 'blog/home.html', {'orders': orders, 'columns': columns, 'col_length':col_length, 'read': read, 'msg':msg})
            else:
                customers = ''
                orders = ''
                form = DateForm()
                msg=''
                return render(request, 'blog/home.html', {'customers': customers, 'orders': orders, 'form': form, 'msg':msg, 'no_data': 'No data found'})

        else:
            customers = ''
            orders = ''
            form = DateForm()
            msg=''
            return render(request, 'blog/home.html', {'customers': customers, 'orders': orders, 'form': form, 'msg':msg})
    else:
        form = DateForm()
        return render(request, 'blog/home.html', {'form': form, 'msg': 'no file exists'})

@login_required
def file_upload(request):
	# context = {
	# 	'posts': Post.objects.all()
	# }
	# return render(request, 'blog/home.html', context)

    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["csv_file"]
            wb = openpyxl.load_workbook(excel_file)

            # getting a particular sheet by name out of many sheets
            worksheet = wb["Sheet1"]
            print(worksheet)

            excel_data = list()
            # iterating over the rows and
            # getting value from each cell in row
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

            form.save()
            return redirect('blog-home')
        msg = 'Please upload file with .csv extension!'
    else:
        msg = ''
        form = DocumentForm()
    return render(request, 'blog/file_upload.html', {
        'form': form, 'msg': msg
    })

@login_required
def upload_csv(request):
    data = {}
    if "GET" == request.method:
        print('simple')
        return render(request, "blog/upload_csv.html", data)
    # if not GET, then proceed
    try:
        csv_file = request.FILES["csv_file"]
        if not csv_file.name.endswith('.csv'):
            print('not csv file ios here')
            messages.error(request,'File is not CSV type')
            return HttpResponseRedirect(reverse("upload_csv"), {'msg':'Please upload file with .csv extension!'})
        #if file is too large, return
        # if csv_file.multiple_chunks():
        #     print('file is too big')
        #     messages.error(request,"Uploaded file is too big (%.2f MB)." % (csv_file.size/(1000*1000),))
        #     return HttpResponseRedirect(reverse("upload_csv"), {'msg':'file is too big'})


        file_data = csv_file.read().decode("utf-8")
        lines = file_data.split("\n")
        print('lines are:', lines)
        del lines[0]
        print('pop lines',lines)
        # loop over the lines and save them in db. If error , store as string and then display
        for line in lines:
            if line:
                # print('line 1',line)
                fields = line.split(",")
                # print('SDKNSDJKNSDJ',fields)
                # data_dict = {}
                # data_dict["name"] = fields[0]
                # data_dict["start_date_time"] = fields[1]
                # data_dict["end_date_time"] = fields[2]
                # data_dict["notes"] = fields[3]
                # # print('data_dict', data_dict)
                if fields[10] == '':
                    fields[10] = datetime.now()

                if fields[17] == '':
                    fields[17] = datetime.now()

                print('fieldno10',fields[10])
                print('is se agey nahi ja raha why ')
                customer_exists = Customers.objects.filter(customer_id=fields[0])
                if customer_exists.exists():
                    print('is k andar ata hai ya nahi   ')
                    # customer_exists.delete()
                    customer = customer_exists.update(
                        customer_id=fields[0],
                        first_name=fields[1],
                        last_name=fields[2],
                        address_1=fields[3],
                        address_2=fields[4],
                        city=fields[5],
                        state=fields[6],
                        zip_code=fields[7],
                        country=fields[8],
                        country_code=fields[9],
                        origin_date=fields[10],
                        cust_email=fields[11],
                        cust_media_origin=fields[13],
                        cust_notes=fields[14]

                    )
                    print('customer is updated')
                else:
                    customer = Customers.objects.create(
                        customer_id=fields[0],
                        first_name=fields[1],
                        last_name=fields[2],
                        address_1=fields[3],
                        address_2=fields[4],
                        city=fields[5],
                        state=fields[6],
                        zip_code=fields[7],
                        country=fields[8],
                        country_code=fields[9],
                        origin_date=fields[10],
                        cust_email=fields[11],
                        cust_media_origin=fields[13],
                        cust_notes=fields[14]

                    )
                    print('customer is not updated')
                cust = Customers.objects.filter(customer_id=fields[0])
                print('cust:', cust[0])

                if fields[18] == '':
                    fields[18] = 0.00

                if fields[19] == '':
                    fields[19] = 0.00

                if fields[22] == '':
                    fields[22] = 0.00

                print('fileldein', fields[18], fields[19])

                order_exists = Orders.objects.filter(order_id=fields[15])
                if order_exists.exists():
                    # order_exists.delete()
                    order = order_exists.update(
                        customer_id=cust[0],
                        order_id=fields[15],
                        product_id=fields[16],
                        order_date=fields[17],
                        unit_price=fields[18],
                        order_total=fields[19],
                        order_detail=fields[20],
                        refund_amount=fields[22],
                        ord_media_origin=fields[23]
                    )
                    print('order is updated')
                else:
                    order = Orders.objects.create(
                        customer_id=cust[0],
                        order_id=fields[15],
                        product_id=fields[16],
                        order_date=fields[17],
                        unit_price=fields[18],
                        order_total=fields[19],
                        order_detail=fields[20],
                        refund_amount=fields[22],
                        ord_media_origin=fields[23]
                    )
                    print('order is not updated')

                try:
                    form = DocumentForm()
                    if form.is_valid():
                        print('here is the form',form)
                        form.save()
                    else:
                        logging.getLogger("error_logger").error(form.errors.as_json())
                except Exception as e:
                    logging.getLogger("error_logger").error(repr(e))
                    pass

        messages.success(request, 'File is uploaded successfully')

    except Exception as e:
        logging.getLogger("error_logger").error("Unable to upload file. "+repr(e))
        messages.error(request,"Unable to upload file. "+repr(e))

    return HttpResponseRedirect(reverse("blog-home"), {'msg':'file is uploaded successfully'})

# #Show all posts
# class PostListView(ListView):
# 	model = Post
# 	template_name = 'blog/home.html'
# 	context_object_name = 'posts'
# 	# ordering = ['-date_posted'] # the minus sign (-) means to return the posts in reverse chronological order
# 	paginate_by = 5
#
#
# # Show posts from the selected user
# class UserPostListView(ListView):
# 	model = Post
# 	template_name = 'blog/user_posts.html'
# 	context_object_name = 'posts'
# 	ordering = ['-date_posted'] # the minus sign (-) means to return the posts in reverse chronological order
# 	paginate_by = 5
#
# 	def get_queryset(self):
# 		user = get_object_or_404(User, username=self.kwargs.get('username'))
# 		return Post.objects.filter(author=user).order_by('-date_posted')
#
#
# class PostDetailView(DetailView):
# 	model = Post
# 	template_name = 'blog/post_detail.html'
#
#
# class PostCreateView(LoginRequiredMixin, CreateView):
# 	model = Post
# 	fields = ['title', 'content']
#
# 	def form_valid(self, form):
# 		form.instance.author = self.request.user
# 		return super().form_valid(form)
#
#
# class PostUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
# 	model = Post
# 	fields = ['title', 'content']
#
# 	def form_valid(self, form):
# 		form.instance.author = self.request.user
# 		return super().form_valid(form)
#
# 	# Only allow the creator of a post to update it
# 	def test_func(self):
# 		post = self.get_object()
# 		if self.request.user == post.author:
# 			return True
# 		return False
#
#
# class PostDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
# 	model = Post
# 	success_url = '/'   # send the user to the homepage if they delete the post
# 	# Only allow the creator of a post to delete it
# 	def test_func(self):
# 		post = self.get_object()
# 		if self.request.user == post.author:
# 			return True
# 		return False

@login_required
def about(request):
	return render(request, 'blog/about.html', {'title': 'About'})
